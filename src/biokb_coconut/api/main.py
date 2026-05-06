import logging
import operator
import os
import secrets
from ast import stmt
from contextlib import asynccontextmanager
from io import BytesIO
from turtle import st
from typing import AsyncGenerator, Generator, Sequence, Tuple

import pandas as pd
import uvicorn
from fastapi import Depends, FastAPI, HTTPException, Query, status
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, Response, StreamingResponse
from fastapi.security import HTTPBasic, HTTPBasicCredentials
from rdkit import Chem
from rdkit.Chem import AllChem, Draw, rdDepictor
from sqlalchemy import Engine, create_engine, func, select
from sqlalchemy.engine.row import Row
from sqlalchemy.orm import Session
from sqlalchemy.sql import operators

from biokb_coconut.api import schemas
from biokb_coconut.api.query_tools import (
    SASearchResults,
    build_dynamic_query,
    create_dynamic_query_filters,
)
from biokb_coconut.api.tags import Tag
from biokb_coconut.constants import (
    DB_DEFAULT_CONNECTION_STR,
    NEO4J_PASSWORD,
    NEO4J_URI,
    NEO4J_USER,
    ZIPPED_TTLS_PATH,
)
from biokb_coconut.db import manager, models
from biokb_coconut.rdf.neo4j_importer import Neo4jImporter
from biokb_coconut.rdf.turtle import TurtleCreator

# Configure logging
logging.basicConfig(
    level=logging.INFO, format="%(asctime)s - %(name)s - %(levelname)s - %(message)s"
)
logger = logging.getLogger(__name__)

USERNAME = os.environ.get("COCONUT_API_USERNAME", "admin")
PASSWORD = os.environ.get("COCONUT_API_PASSWORD", "admin")


def get_engine() -> Engine:
    conn_url = os.environ.get("CONNECTION_STR", DB_DEFAULT_CONNECTION_STR)
    engine: Engine = create_engine(conn_url)
    return engine


def get_session() -> Generator[Session, None, None]:
    engine: Engine = get_engine()
    session = Session(bind=engine)
    try:
        yield session
    finally:
        session.close()


@asynccontextmanager
async def lifespan(app: FastAPI) -> AsyncGenerator[None, None]:
    """Initialize app resources on startup and cleanup on shutdown."""
    engine = get_engine()
    manager.DbManager(engine)
    yield
    # Clean up resources if needed
    pass


description = (
    """A RESTful API for Coconut. Reference: https://coconut.naturalproducts.net/"""
)

app = FastAPI(
    title="RESTful API for Coconut",
    description=description,
    version="0.1.0",
    lifespan=lifespan,
    # Don't set root_path since the reverse proxy doesn't strip the prefix
    # root_path should only be used when a proxy removes a path prefix before forwarding
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Allows all origins
    allow_credentials=True,
    allow_methods=["*"],  # Allows all methods
    allow_headers=["*"],  # Allows all headers
)


def run_api(host: str = "0.0.0.0", port: int = 8000) -> None:
    uvicorn.run(
        app="biokb_coconut.api.main:app",
        host=host,
        port=port,
        log_level="warning",
    )


def verify_credentials(
    credentials: HTTPBasicCredentials = Depends(HTTPBasic()),
) -> None:
    is_correct_username = secrets.compare_digest(credentials.username, USERNAME)
    is_correct_password = secrets.compare_digest(credentials.password, PASSWORD)
    if not (is_correct_username and is_correct_password):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Invalid credentials",
            headers={"WWW-Authenticate": "Basic"},
        )


###############################################################################
# Database Management
###############################################################################


@app.post(
    path="/import_data/",
    response_model=dict[str, int],
    tags=[Tag.DBMANAGE],
)
def import_data(
    credentials: HTTPBasicCredentials = Depends(verify_credentials),
    force_download: bool = Query(
        False,
        description=(
            "Whether to re-download data files even if they already exist,"
            " ensuring the newest version."
        ),
    ),
    delete_files: bool = Query(
        False,
        description=(
            "Whether to delete the downloaded files"
            " after importing them into the database."
        ),
    ),
) -> dict[str, int]:
    """Download data (if not exists) and load in database.

    Can take up to 15 minutes to complete.
    """
    try:
        dbm = manager.DbManager()
        result = dbm.import_data(
            force_download=force_download, delete_files=delete_files
        )
    except Exception as e:
        logger.error(f"Error importing data: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error importing data. {e}",
        ) from e
    return result


@app.get("/export_ttls/", tags=[Tag.DBMANAGE])
async def create_ttls(
    credentials: HTTPBasicCredentials = Depends(verify_credentials),
    force_create: bool = Query(
        False,
        description="Whether to re-generate the TTL files even if they already exist.",
    ),
) -> FileResponse:

    file_path = ZIPPED_TTLS_PATH
    if not os.path.exists(file_path) or force_create:
        try:
            TurtleCreator().create_ttls()
        except Exception as e:
            logger.error(f"Error generating TTL files: {e}")
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="Error generating TTL files. Data already imported?",
            ) from e
    return FileResponse(
        path=file_path, filename="coconut_ttls.zip", media_type="application/zip"
    )


@app.get("/import_neo4j/", tags=[Tag.DBMANAGE])
async def import_neo4j(
    credentials: HTTPBasicCredentials = Depends(verify_credentials),
    uri: str | None = Query(
        default=os.environ.get("NEO4J_URI", NEO4J_URI),
        description="The Neo4j URI. If not provided, "
        "the default from environment variable is used.",
    ),
    user: str | None = Query(
        default=os.environ.get("NEO4J_USER", NEO4J_USER),
        description="The Neo4j user. If not provided,"
        " the default from environment variable is used.",
    ),
    password: str | None = Query(
        NEO4J_PASSWORD,
        description="The Neo4j password. If not provided,"
        " the default from environment variable is used.",
    ),
) -> dict[str, str]:
    """Import RDF turtle files in Neo4j."""
    try:
        if not os.path.exists(ZIPPED_TTLS_PATH):
            raise HTTPException(
                status_code=status.HTTP_405_METHOD_NOT_ALLOWED,
                detail=(
                    "Zipped TTL files not found. Please "
                    "generate them first using /export_ttls/ endpoint."
                ),
            )
        importer = Neo4jImporter(neo4j_uri=uri, neo4j_user=user, neo4j_pwd=password)
        importer.import_ttls()
    except Exception as e:
        logger.error(f"Error importing data into Neo4j: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error importing data into Neo4j: {e}",
        ) from e
    return {"status": "Neo4j import completed successfully."}


###############################################################################
# Compound
###############################################################################


@app.get(
    "/compounds/",
    response_model=schemas.CompoundSearchResult,
    tags=[Tag.COMPOUND],
)
async def search_compounds(
    search: schemas.CompoundOrganismSearch = Depends(),
    session: Session = Depends(get_session),
):
    """
    Search compounds. Returns a list of compounds with their DOIs,
    synonyms, organisms, collections, and CAS numbers.
    """
    try:
        filters = create_dynamic_query_filters(search, models.Compound)
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_102_PROCESSING,
            detail=e,
        )

    # Build the count_stmt with the same filters to get the total count
    # -----------------------------------------------------------------

    # Add organism filter if organism_name is provided
    if search.organism_name:
        filters.append(models.Organism.name.like(search.organism_name))

    count_stmt = select(func.count(models.Compound.id)).select_from(models.Compound)

    # Only join with organisms if organism_name filter is provided to avoid unnecessary joins
    if search.organism_name:
        count_stmt = count_stmt.join(models.Compound.organisms).group_by(
            models.Compound.id
        )

    count_stmt = count_stmt.where(*filters)
    # Because of the potential group_by when filtering by organism, we need to wrap
    # the count_stmt in another select to get the total count
    if search.organism_name:
        count_stmt = select(func.count()).select_from(count_stmt.subquery())

    # logger.info(f"Executing count query: {count_stmt}")
    count = session.execute(count_stmt).scalar() or 0

    # Build the main query with the same filters

    query = session.query(
        models.Compound,
    ).select_from(models.Compound)
    if search.organism_name:
        query = query.outerjoin(models.CompoundOrganism).outerjoin(models.Organism)
    query = query.where(*filters)

    # NOTE: This assumes that the 'order_by' field corresponds to a column in the Compound model.
    if search.order_by is not None and not hasattr(models.Compound, search.order_by):
        raise HTTPException(
            status_code=400,
            detail=f"Invalid 'order_by' field: '{search.order_by}'. No such column in the model.",
        )
    if search.order_by is not None:
        if search.order_desc == "true" or search.order_desc is True:
            query = query.order_by(getattr(models.Compound, search.order_by).desc())
        else:
            query = query.order_by(getattr(models.Compound, search.order_by).asc())

    if search.limit is not None:
        query = query.limit(search.limit)
    if search.offset is not None:
        query = query.offset(search.offset)

    # logger.info(f"Executing query: {query}")

    return {
        "count": count,
        "limit": search.limit,
        "offset": search.offset,
        "results": query.all(),
    }


@app.get(
    "/organism/suggestions",
    response_model=list[str],
    tags=[Tag.ORGANISM],
)
async def suggest_organisms(
    organism_search: str = Query(..., description="Organism name to search for"),
    session: Session = Depends(get_session),
) -> Sequence[str]:
    """
    Search compounds. Returns a list of compounds with their DOIs,
    synonyms, organisms, collections, and CAS numbers.
    """
    stmt = (
        select(models.Organism.name)
        .where(models.Organism.name.ilike(f"{organism_search}%"))
        .limit(10)
        .order_by(models.Organism.name.asc())
    )
    result: Sequence[str] = session.execute(stmt).scalars().all()
    return result


# @app.get(
#     "/compounds/", response_model=schemas.CompoundSearchResult, tags=[Tag.COMPOUND]
# )
# async def search_compounds(
#     search: schemas.CompoundSearch = Depends(),
#     session: Session = Depends(get_session),
# ) -> SASearchResults | dict[str, str]:
#     """
#     Search compounds. Returns a list of compounds with their DOIs,
#     synonyms, organisms, collections, and CAS numbers.
#     """
#     return build_dynamic_query(
#         search_obj=search,
#         model_cls=models.Compound,
#         db=session,
#     )


@app.get(
    "/compounds/export",
    response_model=schemas.CompoundSearchResult,
    tags=[Tag.COMPOUND],
)
async def export_compounds(
    search: schemas.CompoundSearchExportFile = Depends(),
    session: Session = Depends(get_session),
) -> StreamingResponse:
    """
    Export compounds. Returns a list of max. 1000 compounds with their organisms.
    """
    c = models.Compound
    filters = create_dynamic_query_filters(search_obj=search, model_cls=c)
    # get all which are not foreign keys
    columns = [
        col
        for col in c.__table__.columns
        if not col.foreign_keys and col.primary_key is False
    ]

    stmt_compounds = (
        select(
            *columns,
            models.ChemicalClass.name.label("chemical_class"),
            models.ChemicalSubClass.name.label("chemical_sub_class"),
            models.DirectParentClassification.name.label(
                "direct_parent_classification"
            ),
            models.ChemicalSuperClass.name.label("chemical_super_class"),
            models.NpClassifierPathway.name.label("np_classifier_pathway"),
            models.NpClassifierClass.name.label("np_classifier_class"),
            models.NpClassifierSuperclass.name.label("np_classifier_superclass"),
            func.group_concat(models.Organism.name.distinct()).label("organism_names"),
        )
        .select_from(c)
        .outerjoin(models.CompoundOrganism)
        .outerjoin(models.Organism)
        .outerjoin(models.ChemicalClass)
        .outerjoin(models.ChemicalSubClass)
        .outerjoin(models.DirectParentClassification)
        .outerjoin(models.ChemicalSuperClass)
        .outerjoin(models.NpClassifierPathway)
        .outerjoin(models.NpClassifierClass)
        .outerjoin(models.NpClassifierSuperclass)
        .where(*filters)
        .group_by(c.identifier)
        .limit(1000)
    )

    if search.order_by:
        order_col = getattr(c, search.order_by, None)
        if order_col is not None:
            if search.order_desc:
                stmt_compounds = stmt_compounds.order_by(order_col.desc())
            else:
                stmt_compounds = stmt_compounds.order_by(order_col.asc())

    result_compounds = session.execute(stmt_compounds).all()
    df_compounds = pd.DataFrame(result_compounds)

    OP_SYMBOLS = {
        operator.gt: ">",
        operator.lt: "<",
        operator.ge: ">=",
        operator.le: "<=",
        operator.eq: "=",
        operator.ne: "!=",
        operators.between_op: "BETWEEN",
        operators.like_op: "LIKE",
        operators.is_: "IS",
    }

    col_replace = {}
    for filter in filters:
        operator_symbol = OP_SYMBOLS.get(filter.operator, str(filter.operator))
        column_name = getattr(filter.left, "key", None) or getattr(
            filter.left, "name", None
        )
        if column_name is None:
            logger.warning(
                f"Could not extract column name from filter: {filter}. Skipping this filter in column renaming."
            )
            continue
        column_name = (
            column_name.replace("_id", "")
            if column_name.endswith("_id")
            else column_name
        )
        value = (
            str(filter.right.value)
            if hasattr(filter.right, "value")
            else str(filter.right)
        )
        col_replace[column_name] = f"{column_name} {operator_symbol} {value}"
    df_compounds.rename(columns=col_replace, inplace=True)

    # organism

    output = BytesIO()
    # create Excel file in memory with multiple sheets
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_compounds.to_excel(writer, index=False, sheet_name="compounds")

        # Apply light orange fill to the 'identifier' column header in 'compounds' sheet
        from openpyxl.styles import PatternFill

        ws_comp = writer.sheets["compounds"]

        # freeze the header row and the first column
        ws_comp.freeze_panes = "B2"

        # header bold
        for cell in ws_comp[1]:
            cell.font = cell.font.copy(bold=True)
            cell.fill = PatternFill(start_color="faffa4", fill_type="solid")

        lable_color = PatternFill(start_color="3eff13", fill_type="solid")
        for col in col_replace.values():
            if col in df_compounds.columns:
                col_index = df_compounds.columns.get_loc(col)
                if isinstance(col_index, int):
                    col_idx = col_index + 1  # 1-based index
                    ws_comp.cell(row=1, column=col_idx).fill = lable_color

    # Move cursor to beginning
    output.seek(0)

    # Return as streaming response
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=ethcstwin_selected.xlsx"},
    )


@app.get(
    "/compounds/statistics/",
    response_model=schemas.CompoundSearchResultStatistics,
    tags=[Tag.COMPOUND],
)
async def get_compounds_statistics(
    search: schemas.CompoundOrganismSearch = Depends(),
    session: Session = Depends(get_session),
) -> schemas.CompoundSearchResultStatistics:
    """Get statistics of compounds matching the search criteria, including count and quartiles of various properties.

    This will only work for the first 10,000 results to avoid performance issues."""
    c = models.Compound
    filters = create_dynamic_query_filters(search_obj=search, model_cls=c)
    if search.organism_name:
        filters.append(models.Organism.name.like(search.organism_name))

    stmt = select(
        c.total_atom_count,
        c.heavy_atom_count,
        c.molecular_weight,
        c.alogp,
        c.topological_polar_surface_area,
        c.rotatable_bond_count,
        c.hydrogen_bond_acceptors,
        c.hydrogen_bond_donors,
        c.hydrogen_bond_acceptors_lipinski,
        c.hydrogen_bond_donors_lipinski,
        c.aromatic_rings_count,
        c.qed_drug_likeliness,
        c.formal_charge,
        c.fractioncsp3,
        c.number_of_minimal_rings,
        c.van_der_walls_volume,
        c.np_likeness,
        c.contains_sugar,
        c.contains_ring_sugars,
        c.contains_linear_sugars,
        c.np_classifier_is_glycoside,
        c.number_of_organisms,
    ).where(*filters)
    if search.organism_name:
        stmt = stmt.join(models.CompoundOrganism).join(models.Organism).group_by(c.id)

    result = session.execute(stmt).all()
    df = pd.DataFrame(result).astype(float)
    all = df.shape[0]
    statistics = {}
    for col in df.columns:
        statistics[col] = {
            "min": round(df[col].min(), 2),
            "q25": round(df[col].quantile(0.25), 2),
            "q50": round(df[col].quantile(0.5), 2),
            "q75": round(df[col].quantile(0.75), 2),
            "max": round(df[col].max(), 2),
            "not_null_percentage": round(int(df[col].notnull().sum()) / all * 100, 1),
        }

    # statistics for boolean properties
    boolean_cols = [
        "contains_sugar",
        "contains_ring_sugars",
        "contains_linear_sugars",
        "np_classifier_is_glycoside",
    ]
    for col in boolean_cols:
        true_count = df[col].sum()
        false_count = (df[col] == False).sum()  # noqa: E712
        null_count = df[col].isnull().sum()
        statistics[col] = {
            "true_percentage": round(true_count / all * 100, 1),
            "false_percentage": round(false_count / all * 100, 1),
            "null_percentage": round(null_count / all * 100, 1),
        }

    return schemas.CompoundSearchResultStatistics(**statistics)


@app.get("/compound/", response_model=schemas.CompoundDetail, tags=[Tag.COMPOUND])
async def get_compound(
    session: Session = Depends(get_session),
    identifier: str = Query(
        ..., description="Compound identifier", examples=["CNP0581134.2"]
    ),
) -> models.Compound | None:
    """
    Search compounds. Returns a list of compounds with their DOIs,
    synonyms, organisms, collections, and CAS numbers.
    """
    return (
        session.query(models.Compound)
        .where(models.Compound.identifier == identifier)
        .first()
    )


@app.get(
    "/compound/molfile/{identifier}",
    response_model=str,
    tags=[Tag.COMPOUND],
)
async def get_compound_molfile(
    session: Session = Depends(get_session), identifier: str | None = None
):
    """Get a 3D molfile string by compound identifier."""
    standard_inchi: str | None = (
        session.query(models.Compound.standard_inchi)
        .where(models.Compound.identifier == identifier)
        .scalar()
    )
    if standard_inchi is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Compound with identifier {identifier} not found.",
        )
    mol = Chem.MolFromInchi(standard_inchi)
    if mol is None:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=(
                f"Failed to parse standard InChI for compound identifier {identifier}."
            ),
        )
    # Add hydrogens (important for 3D geometry)
    mol = Chem.AddHs(mol)

    # Generate 3D coordinates
    res = AllChem.EmbedMolecule(mol, AllChem.ETKDG())  # type: ignore

    if res != 0:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=f"RDKit failed to generate 3D coordinates for compound identifier {identifier}.",
        )

    # Optimize geometry
    AllChem.UFFOptimizeMolecule(mol)  # type: ignore
    return Chem.MolToMolBlock(mol)


@app.get("/compound/image/{identifier}/", response_class=Response, tags=[Tag.COMPOUND])
async def get_compound_image_by_id(
    identifier: str,
    session: Session = Depends(get_session),
    width: int = Query(400, ge=64, le=2048, description="Output image width in px"),
    height: int = Query(300, ge=64, le=2048, description="Output image height in px"),
) -> Response:
    """Render a PNG image for a compound by DB primary key via its standard InChI.

    The InChI is converted to an RDKit molecule, transformed to a Molfile in memory,
    and then rendered to PNG.
    """
    standard_inchi = (
        session.query(models.Compound.standard_inchi)
        .where(models.Compound.identifier == identifier)
        .scalar()
    )

    if standard_inchi is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Compound with identifier {identifier} not found.",
        )
    if not standard_inchi.strip():
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=f"Compound with identifier {identifier} has no standard InChI.",
        )

    mol = Chem.MolFromInchi(standard_inchi)
    if mol is None:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=(
                f"Failed to parse standard InChI for compound identifier {identifier}."
            ),
        )

    # Build a Molfile explicitly from the InChI-derived molecule before drawing.
    rdDepictor.Compute2DCoords(mol)
    molfile = Chem.MolToMolBlock(mol)
    mol_from_molfile = Chem.MolFromMolBlock(molfile)
    if mol_from_molfile is None:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=(
                f"Failed to create Molfile representation for compound identifier {identifier}."
            ),
        )

    image = Draw.MolToImage(mol_from_molfile, size=(width, height))
    image_buffer = BytesIO()
    image.save(image_buffer, format="PNG")

    return Response(content=image_buffer.getvalue(), media_type="image/png")


@app.get("/compound/name/suggestions", response_model=list[str], tags=[Tag.COMPOUND])
async def get_compound_name_suggestions(
    session: Session = Depends(get_session),
    name: str = Query(
        ..., description="Compound name", examples=["Aspirin"], min_length=3
    ),
) -> list[str]:
    """
    Search compounds by name. Returns a list of compound names.

    The name query is case-insensitive and supports partial matches. For example, a query of "asp" may return "Aspirin", "Aspartame", etc.
    """
    stmt = (
        select(models.Compound.name)
        .where(models.Compound.name.ilike(f"{name}%"))
        .limit(10)
        .order_by(models.Compound.name.asc())
    )
    result = session.execute(stmt).all()
    return [row[0] for row in result] if result else []


@app.get("/dois/", response_model=schemas.DOISearchResult, tags=[Tag.COMPOUND])
async def search_dois(
    search: schemas.DOISearch = Depends(),
    session: Session = Depends(get_session),
) -> SASearchResults | dict[str, str]:
    """
    Search DOIs. Returns a list of DOIs with their compounds.
    """
    return build_dynamic_query(
        search_obj=search,
        model_cls=models.DOI,
        db=session,
    )


@app.get(
    "/organisms/", response_model=schemas.OrganismSearchResult, tags=[Tag.COMPOUND]
)
async def search_organisms(
    search: schemas.OrganismSearch = Depends(),
    session: Session = Depends(get_session),
) -> SASearchResults | dict[str, str]:
    """
    Search organisms. Returns a list of organisms with their compounds.
    """
    return build_dynamic_query(
        search_obj=search,
        model_cls=models.Organism,
        db=session,
    )


@app.get("/synonyms/", response_model=schemas.SynonymSearchResult, tags=[Tag.COMPOUND])
async def search_synonyms(
    search: schemas.SynonymSearch = Depends(),
    session: Session = Depends(get_session),
) -> SASearchResults | dict[str, str]:
    """
    Search synonyms. Returns a list of synonyms with their compounds.
    """
    return build_dynamic_query(
        search_obj=search,
        model_cls=models.Synonym,
        db=session,
    )


@app.get(
    "/collections/", response_model=schemas.CollectionSearchResult, tags=[Tag.COMPOUND]
)
async def search_collections(
    search: schemas.CollectionSearch = Depends(),
    session: Session = Depends(get_session),
) -> SASearchResults | dict[str, str]:
    """
    Search collections. Returns a list of collections with their compound identifiers.
    """
    return build_dynamic_query(
        search_obj=search,
        model_cls=models.Collection,
        db=session,
    )


@app.get("/collections/names", response_model=list[schemas.Name], tags=[Tag.COMPOUND])
async def get_collection_names(
    session: Session = Depends(get_session),
    id: int | None = Query(
        None, description="Optional collection ID to filter results"
    ),
    name: str | None = Query(
        None, description="Optional collection name to filter results"
    ),
) -> Sequence[Row[Tuple[int, str]]]:
    """
    Returns a list of collection names.
    """
    stmt = select(models.Collection.id, models.Collection.name)
    if id:
        stmt = stmt.where(models.Collection.id == id)
    if not id and name:
        stmt = stmt.where(models.Collection.name.ilike(name))
    result: Sequence[Row[Tuple[int, str]]] = session.execute(stmt).all()
    return result


@app.get("/cas/", response_model=schemas.CASSearchResult, tags=[Tag.COMPOUND])
async def search_cas(
    search: schemas.CASSearch = Depends(),
    session: Session = Depends(get_session),
) -> SASearchResults | dict[str, str]:
    """
    Search CAS numbers. Returns a list of CAS numbers with their compounds.
    """
    return build_dynamic_query(
        search_obj=search,
        model_cls=models.CAS,
        db=session,
    )


@app.get(
    "/chemical_class/",
    response_model=schemas.ChemicalClassSearchResult,
    tags=[Tag.COMPOUND],
)
async def search_chemical_class(
    search: schemas.ChemicalClassSearch = Depends(),
    session: Session = Depends(get_session),
) -> SASearchResults | dict[str, str]:
    """
    Search CAS numbers. Returns a list of CAS numbers with their compounds.
    """
    return build_dynamic_query(
        search_obj=search,
        model_cls=models.ChemicalClass,
        db=session,
    )


@app.get(
    "/chemical_class/names", response_model=list[schemas.Name], tags=[Tag.COMPOUND]
)
async def get_chemical_class_names(
    session: Session = Depends(get_session),
    id: int | None = Query(
        None, description="Optional chemical class ID to filter results"
    ),
    name: str | None = Query(
        None, description="Optional chemical class name to filter results"
    ),
) -> Sequence[Row[Tuple[int, str]]]:
    """
    Returns a list of chemical class names.
    """
    stmt = select(models.ChemicalClass.id, models.ChemicalClass.name)
    if id:
        stmt = stmt.where(models.ChemicalClass.id == id)
    if not id and name:
        stmt = stmt.where(models.ChemicalClass.name.ilike(name))
    result: Sequence[Row[Tuple[int, str]]] = session.execute(stmt).all()
    return result


@app.get(
    "/chemical_sub_class/",
    response_model=schemas.ChemicalSubClassSearchResult,
    tags=[Tag.COMPOUND],
)
async def search_chemical_sub_class(
    search: schemas.ChemicalSubClassSearch = Depends(),
    session: Session = Depends(get_session),
) -> SASearchResults | dict[str, str]:
    """
    Search chemical sub classes. Returns a list of chemical sub classes
    with their compounds.
    """
    return build_dynamic_query(
        search_obj=search,
        model_cls=models.ChemicalSubClass,
        db=session,
    )


@app.get(
    "/chemical_sub_class/names", response_model=list[schemas.Name], tags=[Tag.COMPOUND]
)
async def get_chemical_sub_class_names(
    session: Session = Depends(get_session),
    id: int | None = Query(
        None, description="Optional chemical sub class ID to filter results"
    ),
    name: str | None = Query(
        None, description="Optional chemical sub class name to filter results"
    ),
) -> Sequence[Row[Tuple[int, str]]]:
    """
    Returns a list of chemical sub class names.
    """
    stmt = select(models.ChemicalSubClass.id, models.ChemicalSubClass.name)
    if id:
        stmt = stmt.where(models.ChemicalSubClass.id == id)
    if not id and name:
        stmt = stmt.where(models.ChemicalSubClass.name.ilike(name))
    result: Sequence[Row[Tuple[int, str]]] = session.execute(stmt).all()
    return result


@app.get(
    "/direct_parent_classification/",
    response_model=schemas.DirectParentClassificationSearchResult,
    tags=[Tag.COMPOUND],
)
async def search_direct_parent_classification(
    search: schemas.DirectParentClassificationSearch = Depends(),
    session: Session = Depends(get_session),
) -> SASearchResults | dict[str, str]:
    """
    Search direct parent classifications. Returns a list of direct parent
    classifications with their compounds.
    """
    return build_dynamic_query(
        search_obj=search,
        model_cls=models.DirectParentClassification,
        db=session,
    )


@app.get(
    "/direct_parent_classification/names",
    response_model=list[schemas.Name],
    tags=[Tag.COMPOUND],
)
async def get_direct_parent_classification_names(
    session: Session = Depends(get_session),
    id: int | None = Query(
        None, description="Optional direct parent classification ID to filter results"
    ),
    name: str | None = Query(
        None, description="Optional direct parent classification name to filter results"
    ),
) -> Sequence[Row[Tuple[int, str]]]:
    """
    Returns a list of direct parent classification names.
    """

    stmt = select(
        models.DirectParentClassification.id, models.DirectParentClassification.name
    )
    if id:
        stmt = stmt.where(models.DirectParentClassification.id == id)
    if not id and name:
        stmt = stmt.where(models.DirectParentClassification.name.ilike(name))
    result: Sequence[Row[Tuple[int, str]]] = session.execute(stmt).all()
    return result


@app.get(
    "/chemical_super_class/",
    response_model=schemas.ChemicalSuperClassSearchResult,
    tags=[Tag.COMPOUND],
)
async def search_chemical_super_class(
    search: schemas.ChemicalSuperClassSearch = Depends(),
    session: Session = Depends(get_session),
) -> SASearchResults | dict[str, str]:
    """
    Search chemical super classes. Returns a list of chemical super classes
    with their compounds.
    """
    return build_dynamic_query(
        search_obj=search,
        model_cls=models.ChemicalSuperClass,
        db=session,
    )


@app.get(
    "/chemical_super_class/names",
    response_model=list[schemas.Name],
    tags=[Tag.COMPOUND],
)
async def get_chemical_super_class_names(
    session: Session = Depends(get_session),
    id: int | None = Query(
        None, description="Optional chemical super class ID to filter results"
    ),
    name: str | None = Query(
        None, description="Optional chemical super class name to filter results"
    ),
) -> Sequence[Row[Tuple[int, str]]]:
    """
    Returns a list of chemical super class names.
    """

    stmt = select(models.ChemicalSuperClass.id, models.ChemicalSuperClass.name)
    if id:
        stmt = stmt.where(models.ChemicalSuperClass.id == id)
    if not id and name:
        stmt = stmt.where(models.ChemicalSuperClass.name.ilike(name))
    result: Sequence[Row[Tuple[int, str]]] = session.execute(stmt).all()
    return result


@app.get(
    "/np_classifier_pathway/",
    response_model=schemas.NpClassifierPathwaySearchResult,
    tags=[Tag.NP_CLASSIFIER],
)
async def search_np_classifier_pathway(
    search: schemas.NpClassifierPathwaySearch = Depends(),
    session: Session = Depends(get_session),
) -> SASearchResults | dict[str, str]:
    """
    Search NP classifier pathways. Returns a list of NP classifier
    pathways with their compounds.
    """
    return build_dynamic_query(
        search_obj=search,
        model_cls=models.NpClassifierPathway,
        db=session,
    )


@app.get(
    "/np_classifier_pathway/names",
    response_model=list[schemas.Name],
    tags=[Tag.NP_CLASSIFIER],
)
async def get_np_classifier_pathway_names(
    session: Session = Depends(get_session),
    id: int | None = Query(
        None, description="Optional NP classifier pathway ID to filter results"
    ),
    name: str | None = Query(
        None, description="Optional NP classifier pathway name to filter results"
    ),
) -> Sequence[Row[Tuple[int, str]]]:
    """
    Returns a list of NP classifier pathway names.
    """

    stmt = select(models.NpClassifierPathway.id, models.NpClassifierPathway.name)
    if id:
        stmt = stmt.where(models.NpClassifierPathway.id == id)
    if not id and name:
        stmt = stmt.where(models.NpClassifierPathway.name.ilike(name))
    result: Sequence[Row[Tuple[int, str]]] = session.execute(stmt).all()
    return result


@app.get(
    "/np_classifier_superclass/",
    response_model=schemas.NpClassifierSuperclassSearchResult,
    tags=[Tag.NP_CLASSIFIER],
)
async def search_np_classifier_superclass(
    search: schemas.NpClassifierSuperclassSearch = Depends(),
    session: Session = Depends(get_session),
) -> SASearchResults | dict[str, str]:
    """
    Search NP classifier superclasses. Returns a list of NP classifier superclasses with their compounds.
    """
    return build_dynamic_query(
        search_obj=search,
        model_cls=models.NpClassifierSuperclass,
        db=session,
    )


@app.get(
    "/np_classifier_superclass/names",
    response_model=list[schemas.Name],
    tags=[Tag.NP_CLASSIFIER],
)
async def get_np_classifier_superclass_names(
    session: Session = Depends(get_session),
    id: int | None = Query(
        None, description="Optional NP classifier superclass ID to filter results"
    ),
    name: str | None = Query(
        None, description="Optional NP classifier superclass name to filter results"
    ),
) -> Sequence[Row[Tuple[int, str]]]:
    """
    Returns a list of NP classifier superclass names.
    """

    stmt = select(models.NpClassifierSuperclass.id, models.NpClassifierSuperclass.name)
    if id:
        stmt = stmt.where(models.NpClassifierSuperclass.id == id)
    if not id and name:
        stmt = stmt.where(models.NpClassifierSuperclass.name.ilike(name))
    result: Sequence[Row[Tuple[int, str]]] = session.execute(stmt).all()
    return result


@app.get(
    "/np_classifier_class/",
    response_model=schemas.NpClassifierClassSearchResult,
    tags=[Tag.NP_CLASSIFIER],
)
async def search_np_classifier_class(
    search: schemas.NpClassifierClassSearch = Depends(),
    session: Session = Depends(get_session),
) -> SASearchResults | dict[str, str]:
    """
    Search NP classifier classes. Returns a list of NP classifier classes with their compounds.
    """
    return build_dynamic_query(
        search_obj=search,
        model_cls=models.NpClassifierClass,
        db=session,
    )


@app.get(
    "/np_classifier_class/names",
    response_model=list[schemas.Name],
    tags=[Tag.NP_CLASSIFIER],
)
async def get_np_classifier_class_names(
    session: Session = Depends(get_session),
    id: int | None = Query(
        None, description="Optional NP classifier class ID to filter results"
    ),
    name: str | None = Query(
        None, description="Optional NP classifier class name to filter results"
    ),
) -> Sequence[Row[Tuple[int, str]]]:
    """
    Returns a list of NP classifier class names.
    """

    stmt = select(models.NpClassifierClass.id, models.NpClassifierClass.name)
    if id:
        stmt = stmt.where(models.NpClassifierClass.id == id)
    if not id and name:
        stmt = stmt.where(models.NpClassifierClass.name.ilike(name))
    result: Sequence[Row[Tuple[int, str]]] = session.execute(stmt).all()
    return result
